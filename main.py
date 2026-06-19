"""
Nurse and Tech Scheduling System.
Uses CP-SAT solver to assign staff to shifts based on daily requirements.
"""
import json
import sys
from ortools.sat.python import cp_model
import pandas as pd


from openpyxl.styles import Alignment, PatternFill, Font


def export_to_excel(nurse_data, tech_data, requirements_df, filename="staff_schedule.xlsx"):
    """Exports nurse and tech data to an Excel file."""
    
    def write_table_on_sheet(ws, df, title, start_r, writer, sheet_name, color='CCE5FF'):
        ws.merge_cells(start_row=start_r, start_column=1, end_row=start_r, end_column=df.shape[1])
        cell = ws.cell(row=start_r, column=1)
        cell.value = title
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.font = Font(bold=True)
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_r)
        return start_r + len(df) + 3

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Requirements Sheet
        requirements_df.to_excel(writer, sheet_name='Requirements', index=False, startrow=1)
        ws_req = writer.sheets['Requirements']
        ws_req.merge_cells(start_row=1, start_column=1, end_row=1, end_column=requirements_df.shape[1])
        title_cell = ws_req.cell(row=1, column=1)
        title_cell.value = "Daily Staffing Requirements"
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        title_cell.font = Font(bold=True)

        # Staff Schedule Sheet
        ws_sched = writer.book.create_sheet('Staff Schedule')
        writer.sheets['Staff Schedule'] = ws_sched
        curr_row_sched = 1

        # Nurse Schedule Table
        curr_row_sched = write_table_on_sheet(ws_sched, nurse_data['schedule'], "Nurse Shift Assignments", curr_row_sched, writer, 'Staff Schedule')

        # Tech Schedule Table
        write_table_on_sheet(ws_sched, tech_data['schedule'], "Tech Shift Assignments", curr_row_sched, writer, 'Staff Schedule')

        # Summary Sheet
        ws_sum = writer.book.create_sheet('Summary')
        writer.sheets['Summary'] = ws_sum
        curr_row = 1

        # New Hires Summary
        purple_fill = PatternFill(start_color='E5CCFF', end_color='E5CCFF', fill_type='solid')
        
        ws_sum.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
        title_cell = ws_sum.cell(row=curr_row, column=1)
        title_cell.value = "Required Hires for Schedule"
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = purple_fill
        title_cell.font = Font(bold=True)
        curr_row += 1
        
        ws_sum.cell(row=curr_row, column=1).value = "New Nurse Hires"
        ws_sum.cell(row=curr_row, column=1).fill = purple_fill
        ws_sum.cell(row=curr_row, column=2).value = nurse_data['additional_count']
        ws_sum.cell(row=curr_row, column=2).fill = purple_fill
        
        ws_sum.cell(row=curr_row + 1, column=1).value = "New Tech Hires"
        ws_sum.cell(row=curr_row + 1, column=1).fill = purple_fill
        ws_sum.cell(row=curr_row + 1, column=2).value = tech_data['additional_count']
        ws_sum.cell(row=curr_row + 1, column=2).fill = purple_fill
        
        curr_row += 4
        
        # Nurse Comparison
        curr_row = write_table_on_sheet(ws_sum, nurse_data['comparison'], "Nurse Work Day Comparison", curr_row, writer, 'Summary')

        # Tech Comparison
        curr_row = write_table_on_sheet(ws_sum, tech_data['comparison'], "Tech Work Day Comparison", curr_row, writer, 'Summary')

        # Extra Hires (Combined)
        extra_hires = pd.concat(
            [nurse_data['extra_hires'], tech_data['extra_hires']], ignore_index=True
        )
        curr_row = write_table_on_sheet(ws_sum, extra_hires, "Extra Staff Hires", curr_row, writer, 'Summary')

        # Daily Staffing (Combined)
        nurse_summary = nurse_data['summary'].copy()
        nurse_summary.columns = ["Day", "Required Nurses", "Scheduled Nurses"]
        tech_summary = tech_data['summary'].copy()
        tech_summary.columns = ["Day", "Required Techs", "Scheduled Techs"]

        combined_summary = pd.merge(nurse_summary, tech_summary, on="Day")
        write_table_on_sheet(ws_sum, combined_summary, "Daily Staffing Totals", curr_row, writer, 'Summary')

    print(f"Results exported to {filename}")


def create_schedule_model(required_per_day, base_staff_data, staff_type, additional_count):
    """Creates the CP-SAT model for scheduling."""
    model = cp_model.CpModel()

    # Base Staff information
    work_days = [staff["required_days"] for staff in base_staff_data]
    names = [f"{staff['first_name']} {staff['last_name']}" for staff in base_staff_data]
    base_num = len(base_staff_data)

    # Add additional staff
    if additional_count > 0:
        for i in range(additional_count):
            work_days.append(4)
            names.append(f"Extra {staff_type} {i + 1}")

    staff_range = range(len(work_days))
    days = range(len(required_per_day))

    shifts = {}
    for s in staff_range:
        for d in days:
            shifts[(s, d)] = model.NewBoolVar(f'shift_s{s}_d{d}')

    # Constraint 1: Daily requirement
    for d in days:
        model.Add(sum(shifts[(s, d)] for s in staff_range) == required_per_day[d])

    # Constraint 2: Work days per staff
    for s in staff_range:
        if s < base_num:
            model.Add(sum(shifts[(s, d)] for d in days) == work_days[s])
        else:
            model.Add(sum(shifts[(s, d)] for d in days) <= 4)
            model.Add(sum(shifts[(s, d)] for d in days) >= 1)

    return model, shifts, names, work_days, base_num


def get_nurse_role(s_idx, d_idx, solver, shifts, staff_range):
    """Calculates the specific role for a nurse on a given day."""
    day_staff = [i for i in staff_range if solver.Value(shifts[(i, d_idx)])]
    pos = day_staff.index(s_idx)
    if pos == 0:
        return "Float"
    if pos == 1:
        return "Lunch"
    return f"Room {((pos - 2) // 2) + 1}"


def get_tech_role(s_idx, d_idx, solver, shifts, staff_range, day_config):
    """Calculates the specific role for a tech on a given day."""
    day_staff = [i for i in staff_range if solver.Value(shifts[(i, d_idx)])]
    pos = day_staff.index(s_idx)

    num_or = day_config["num_or_rooms"]
    num_scope = day_config["num_scope_rooms"]

    if pos < num_or:
        return f"Room {pos + 1}"
    if pos < num_or + (num_scope * 2):
        scope_idx = (pos - num_or) // 2
        return f"Scope {scope_idx + 1}"
    return "Float"


def get_staff_stats(s_idx, solver_data, days):
    """Extracts stats for a single staff member."""
    solver = solver_data["solver"]
    shifts = solver_data["shifts"]
    names = solver_data["names"]
    work_days = solver_data["work_days"]
    base_num = solver_data["base_num"]

    is_additional = s_idx >= base_num
    st_type = "Base" if not is_additional else "Extra"
    name = names[s_idx]
    actual_days = sum(solver.Value(shifts[(s_idx, d_idx)]) for d_idx in days)
    required = work_days[s_idx] if not is_additional else "1-4 (Max 4)"
    return st_type, name, actual_days, required


def process_solver_results(solver_data):
    """Processes the results from the CP-SAT solver into DataFrames."""
    solver = solver_data["solver"]
    shifts = solver_data["shifts"]
    staff_type = solver_data["staff_type"]
    required_per_day = solver_data["required_per_day"]
    base_num = solver_data["base_num"]
    schedule_params = solver_data.get("schedule_params")

    staff_range = range(len(solver_data["names"]))
    days = range(len(required_per_day))

    # Prepare DataFrames
    schedule_cols = ["Type", "Name"] + [f"Day {d + 1}" for d in days] + ["Total"]
    schedule_rows = []
    comparison_rows = []
    extra_hires_rows = []

    for s_idx in staff_range:
        stats = get_staff_stats(s_idx, solver_data, days)
        st_type, name, actual_days, required = stats

        row = [st_type, name]
        for d_idx in days:
            if solver.Value(shifts[(s_idx, d_idx)]):
                if staff_type == "Nurse":
                    row.append(get_nurse_role(s_idx, d_idx, solver, shifts, staff_range))
                else:
                    if schedule_params:
                        day_config = schedule_params["days"][d_idx]
                        row.append(get_tech_role(s_idx, d_idx, solver, shifts, staff_range, day_config))
                    else:
                        row.append("X")
            else:
                row.append("-")
        row.append(actual_days)
        schedule_rows.append(row)
        comparison_rows.append([staff_type, st_type, name, required, actual_days])
        if s_idx >= base_num:
            extra_hires_rows.append([name, actual_days])

    summary_rows = []
    for d_idx in days:
        count = sum(solver.Value(shifts[(s_idx, d_idx)]) for s_idx in staff_range)
        summary_rows.append([f"Day {d_idx + 1}", required_per_day[d_idx], count])

    return {
        "schedule": pd.DataFrame(schedule_rows, columns=schedule_cols),
        "comparison": pd.DataFrame(comparison_rows,
                                   columns=["Staff Type", "Type", "Name",
                                            "Required Days", "Scheduled Days"]),
        "extra_hires": pd.DataFrame(extra_hires_rows,
                                    columns=[f"Extra {staff_type} Name", "Scheduled Days"]),
        "summary": pd.DataFrame(summary_rows, columns=["Day", "Required", "Scheduled"]),
        "additional_count": solver_data["additional_count"]
    }


def solve_scheduling(staff_type, required_per_day, base_staff_data, additional_count=0,
                     schedule_params=None):
    """
    Main function to solve the scheduling problem for a given staff type.
    """
    model, shifts, names, work_days, base_num = create_schedule_model(
        required_per_day, base_staff_data, staff_type, additional_count
    )

    solver = cp_model.CpSolver()
    status = solver.Solve(model)

    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return process_solver_results({
            "solver": solver,
            "shifts": shifts,
            "names": names,
            "work_days": work_days,
            "base_num": base_num,
            "staff_type": staff_type,
            "required_per_day": required_per_day,
            "additional_count": additional_count,
            "schedule_params": schedule_params
        })
    return None


def calculate_requirements(schedule_params):
    """Calculates daily nurse and tech requirements."""
    nurse_requirements = []
    tech_requirements = []

    for day_config in schedule_params["days"]:
        # Calculate Nurse requirements per day
        # Formula: (OR rooms * 2) + float nurses + lunch relief nurses
        daily_nurse_req = (day_config["num_or_rooms"] * 2) + \
                          day_config["num_float_nurses"] + \
                          day_config["num_lunch_relief_nurses"]
        nurse_requirements.append(daily_nurse_req)

        num_or_rooms = day_config["num_or_rooms"]
        # Updated constraint: always want a dedicated floating tech (1)
        # and no additional constraint on float tech even if more than 3 rooms.
        daily_tech_req = (num_or_rooms * 1) + \
                         (day_config["num_scope_rooms"] * 2) + 1
        tech_requirements.append(daily_tech_req)
    return nurse_requirements, tech_requirements


def main():
    """Main execution function."""
    try:
        with open('staff.json', 'r', encoding='utf-8') as f:
            staff_data = json.load(f)
        with open('schedule-parameters.json', 'r', encoding='utf-8') as f:
            schedule_params = json.load(f)
    except FileNotFoundError as e:
        print(f"Error: {e.filename} not found.")
        sys.exit(1)

    # Calculate requirements per day from schedule-parameters.json
    nurse_requirements, tech_requirements = calculate_requirements(schedule_params)

    # Solve for Nurses
    nurse_added = 0
    nurse_results = None
    while nurse_added <= 20:
        nurse_results = solve_scheduling("Nurse", nurse_requirements, staff_data["nurses"],
                                         nurse_added, schedule_params)
        if nurse_results:
            break
        nurse_added += 1

    if not nurse_results:
        print("no solution possible for nurses")
        sys.exit(1)

    # Solve for Techs
    tech_added = 0
    tech_results = None
    while tech_added <= 20:
        tech_results = solve_scheduling("Tech", tech_requirements, staff_data["techs"], tech_added,
                                         schedule_params)
        if tech_results:
            break
        tech_added += 1

    if not tech_results:
        print("no solution possible for techs")
        sys.exit(1)

    # Display requirements used
    print("Calculated Daily Requirements:")
    requirements_summary_df = pd.DataFrame({
        "Day": [f"Day {i + 1}" for i in range(len(nurse_requirements))],
        "OR Rooms": [d["num_or_rooms"] for d in schedule_params["days"]],
        "Scope Rooms": [d["num_scope_rooms"] for d in schedule_params["days"]],
        "Nurse Required": nurse_requirements,
        "Tech Required": tech_requirements
    })
    print(requirements_summary_df.to_string(index=False))
    print("")

    # Display results
    print(f"\nNurse Solution found with {nurse_added} extra nurses.")
    print("Nurse Schedule:")
    print(nurse_results["schedule"].to_string(index=False))
    print(f"\nTech Solution found with {tech_added} extra techs.")
    print("Tech Schedule:")
    print(tech_results["schedule"].to_string(index=False))

    print("\nStaff Work Comparison:")
    print("Nurse Comparison:")
    print(nurse_results["comparison"].to_string(index=False))
    print("\nTech Comparison:")
    print(tech_results["comparison"].to_string(index=False))

    export_to_excel(nurse_results, tech_results, requirements_summary_df)


if __name__ == '__main__':
    main()
